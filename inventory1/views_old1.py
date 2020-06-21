import csv, io   #for csv
from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpRequest
from .forms import StockCreateForm, SuppliersCreateForm, StockSearchForm,StockUpdateForm, StockFullSearchForm
from .models import *
from .clean_import_files import *
from django.contrib import messages
import openpyxl
from datetime import datetime
from django.contrib.auth.decorators import permission_required  #for updateing from csv

# Create your views here.
 
 
def homepage_view(request):
    #this is the title for now
    #print (request.user)
  
    today = datetime.now().date()
    context= {
     
        "today": today,
        
    }

    return render(request,"home.html", context) 
"""  """


def delete_items(request,pk):
    #instance = get_object_or_404( id=pk)
    instance = Stock.objects.get(id=pk)
    instance.delete()
    return redirect('/list_items')

# 

def about_view(request):
    title='Welcome to my first Django project'
    #today = datetime.now().date()
    aboutmyself= "My name is Avi Sivan. My email onlyavi@gmail.com"
    context= {
        "title":title,
        "today": today,
        "aboutmyself":aboutmyself
        
    }
    
    return render(request, "about.html",context)


def under_construction(request):
   
    return render(request, "under_construction.html")

def add_newitem_view(request):
    form = StockCreateForm(request.POST or None)
    #z = form['item_name'].value()
    if form.is_valid():
        form.save()
        return redirect('/add_newitem')
    
    context ={
        "form":form,
        "title":"Add New Item",
        
    }
    return render(request, "add_newitem.html", context)

def update_items(request,pk):
    #pk is the object we want to update
    queryset= Stock.objects.get(id=pk)
    form=StockUpdateForm(instance=queryset)
    if request.method=='POST':
        form=StockUpdateForm(request.POST, instance=queryset)
        if form.is_valid():
            form.save()
            return redirect('/list_items')
    context={
        "form":form,
        
    }
    return render(request, 'add_newitem.html',context)       



def list_items_view(request):
    header ='List of items'
    searchtitle= "Search Item"
    form = StockSearchForm(request.POST or None)
    queryset = Stock.objects.all()
    context ={
        "header": header,
        "queryset":queryset,
        "form": form,
                          
    }
    if request.method == 'POST':
        queryset = Stock.objects.filter(#item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         category_name__icontains=form['category_name'].value(),
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
           
    if request.method == 'POST':
        queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         #category_name__icontains=form['category_name'].value(),
                                         #item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
      
    context ={
        "form": form,
        "header": header,
        "queryset": queryset,
        "searchtitle": searchtitle
      
                
    }
    
    return render(request, "list_items.html", context)



def StockFSearch_view(request):
    header ='Full Item Search'
    form = StockFullSearchForm(request.POST or None)
    queryset = Stock.objects.all()
  
    context ={
        "header": header,
        "queryset":queryset,
        "form": form,
                          
    }

        
    # if request.method == 'POST':
    #         # z=  form['item_name'].value()
    #         # zz= Stock.objects.values('item_name')
    #         # var2 = Stock.objects.values_list('item_name')
    #         # print ("z: ",z, ", var2:", var2, "zz: ",zz)
    #         # for i in var2:
    #         #     print (i)
    #         #     if z in i:
    #         #         print ("found ",z , " *****")
            
    #         # if isnumeric(form['quantity_min'].value())== False:
    #         #     print ("missing a number ***")
                
                
    #         queryset = Stock.objects.filter(category_name__icontains=form['category_name'].value(),
    #                                         item_name__icontains=form['item_name'].value(),
    #                                         quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value() ),
    #                                         item_fattal_code__icontains=form['item_fattal_code'].value(),
    #                                         description__icontains=form['description'].value(),
                                        
                                        
             #)
    if request.method == 'POST':
        queryset = Stock.objects.filter(#item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         category_name__icontains=form['category_name'].value(),
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
           
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
    #                                      #category_name__icontains=form['category_name'].value(),
    #                                      #item_name__iexact=form['item_name'].value(),
    #    )
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(item_fattal_code__range=(form['item_fattal_code_begin'].value(),form['item_fattal_code_end'].value()),
                                                     
        
    #     )
    
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value()),
                                                     
        
    #     )
    
    
    if request.method == 'POST':
        queryset = Stock.objects.filter(quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value()),
                                    item_fattal_code__range=(form['item_fattal_code_begin'].value(),form['item_fattal_code_end'].value())                                       
                                                     
        
        )
    
  
       
    context ={
        "form": form,
        "header": header,
        "queryset": queryset,
                
    }
    
    
    return render(request, "full_search.html", context)


def csv_excel_file(request):
    prompt = {
        'order': 'MAKE SURE that this is csv or xlsx file, and the fields are in correct order'

    }
    if request.method == "GET":
        return render (request, 'excel_upload1.html', {})
    file_data= request.FILES['file']
    if not file_data.name.endswith('.csv') and not file_data.name.endswith('.xlsx'):
        messages.error(request, "This is not a valid file format (only csv or excel xlsx")
    elif file_data.name.endswith('.csv'):
        #csv_upload(request,file_data)
        csv_file =request.FILES['file']
        data_set = csv_file.read().decode('UTF-8')
        io_string = io.StringIO(data_set)
        next(io_string) #to skip the first line usually headrs
        for column in csv.reader(io_string, delimiter=',', quotechar="|" ):
            skip=False
            sskip=False
            item_fattal_code_check= column[0]
            clean_excel_item_fattal_code(item_fattal_code_check,skip)

            item_name_check= column[1]
            clean_excel_item_name(item_name_check,skip)
            
            sup_name= column[2]
            clean_excel_supplliers_name(sup_name,sskip)
    # item_name = a.cleaned_data.get('item_name')
        #category_name =self.cleaned_data.get('category_name')
        
            
        
            #clean_excel_item_fattal_code(item_name_check,skip)
            if skip == False:
                
                _, created = Stock.objects.update_or_create(
                    item_fattal_code=column[0],
                    item_name= column[1],

                )
            if sskip==False:
                _, created = SupplierInformation.objects.update_or_create(
                    suppliers_name=column[2]

                ) 
            if skip == True:
                skip = False
            if sskip == True:
                sskip = False
            
        context ={}
        #so it can be change in the future
        
        return render(request,'excel_upload1.html',context)



    elif file_data.name.endswith('.xlsx'):
        #excel_upload(request)
        #excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        
        wb = openpyxl.load_workbook(file_data)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)


        # reading a cell
        print(worksheet["A1"].value)

        return render(request, 'excel_upload.html', {"excel_data":excel_data})
    




    


#@permission_required('admin.can_add_log_entry')
def csv_upload(request,file_data):
    template1="excel_upload1.html"
    prompt = {
        'order': 'MAKE SURE that this is csv or xlsx file, and the fields are in correct order'

    }
    if request.method == "GET":
        return render (request, template1, prompt)

    csv_file =request.FILES['file']
    if not csv_file.name.endswith('.csv'):
         messages.error(request, "this is not csv file")

    data_set = csv_file.read().decode('UTF-8')
    io_string = io.StringIO(data_set)
    next(io_string) #to skip the first line usually headrs
    for column in csv.reader(io_string, delimiter=',', quotechar="|" ):
        skip=False
        sskip=False
        item_fattal_code_check= column[0]
        clean_excel_item_fattal_code(item_fattal_code_check,skip)

        item_name_check= column[1]
        clean_excel_item_name(item_name_check,skip)
        
        sup_name= column[2]
        clean_excel_supplliers_name(sup_name,sskip)
   # item_name = a.cleaned_data.get('item_name')
    #category_name =self.cleaned_data.get('category_name')
    
         
    
        #clean_excel_item_fattal_code(item_name_check,skip)
        if skip == False:
            
            _, created = Stock.objects.update_or_create(
                item_fattal_code=column[0],
                item_name= column[1],

             )
        if sskip==False:
            _, created = SupplierInformation.objects.update_or_create(
                suppliers_name=column[2]

            ) 
        if skip == True:
            skip = False
        if sskip == True:
            sskip = False
        
    context ={}
    #so it can be change in the future
    
    #return render(csv_file,'excel_upload1.html',context)


def excel_upload(request):
    if "GET" == request.method:
         return render(request, 'excel_upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "this is not excel  file")

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)


        # reading a cell
        print(worksheet["A1"].value)

    return render(request, 'excel_upload.html', {"excel_data":excel_data})
    


# def excel_import_example(request):
#     if "GET" == request.method:
#          return render(request, 'excel_upload.html', {})
#     else:
#         excel_file = request.FILES["excel_file"]

#         # you may put validations here to check extension or file size
#         if not excel_file.name.endswith('.xlsx'):
#             messages.error(request, "this is not excel  file")

#         wb = openpyxl.load_workbook(excel_file)

#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#         print(worksheet)

#         # getting all sheets
#         sheets = wb.sheetnames
#         print(sheets)

#         # getting active sheet
#         active_sheet = wb.active
#         print(active_sheet)

#         excel_data = list()
#         # iterating over the rows and
#         # getting value from each cell in row
#         for row in worksheet.iter_rows():
#             row_data = list()
#             for cell in row:
#                 row_data.append(str(cell.value))
#             excel_data.append(row_data)


#         # reading a cell
#         print(worksheet["A1"].value)

#     return render(request, 'excel_upload.html', {"excel_data":excel_data})
    


    