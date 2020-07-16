import csv, io   #for csv
from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpRequest
from .forms import StockCreateForm, SuppliersCreateForm, StockSearchForm,StockUpdateForm, StockFullSearchForm, import_csv, import_excel, StockIssueForm,cleanme,StockIssueForm2

from inventory1.issue_functions import *
from .models import *
from .clean_import_files import *
from django.contrib import messages
import openpyxl
from datetime import datetime
from django.contrib.auth.decorators import permission_required  #for updateing from csv
from django.views.generic import FormView
# Create your views here.
 
 
def homepage_view(request):
    #this is the title for now
    #print (request.user)
    # a=10
    # b=7
    # print (int(a / b))
    today = datetime.now().date()
    context= {
     
        "today": today,
        
    }

    return render(request,"home.html", context) 
"""  """


def delete_items(request,pk):
    #instance = get_object_or_404( id=pk)
    instance = Stock.objects.get(id=pk)
    instance.delete()
    return redirect('/list_items')

# 

def about_view(request):
    aboutmyself='Welcome to my first Django project'
    today = datetime.now().date()
    title= "My name is Avi Sivan. My email onlyavi@gmail.com"
    context= {
        "title":title,
        "today": today,
        "aboutmyself":aboutmyself
        
    }
    
    return render(request, "about.html",context)


def under_construction(request):
   
    return render(request, "under_construction.html")

def add_newitem_view(request):
    form = StockCreateForm(request.POST or None)
    #z = form['item_name'].value()
    if form.is_valid():
        form.save()
        return redirect('/add_newitem')
    
    context ={
        "form":form,
        "title":"Add New Item",
        
    }
    return render(request, "add_newitem.html", context)

def update_items(request,pk):
    #pk is the object we want to update
    queryset= Stock.objects.get(id=pk)
    form=StockUpdateForm(instance=queryset)
    if request.method=='POST':
        form=StockUpdateForm(request.POST, instance=queryset)
        if form.is_valid():
            form.save()
            return redirect('/list_items')
    context={
        "form":form,
        
    }
    return render(request, 'add_newitem.html',context)    

def item_issue_view(request):
    header ='Full Item Search'
    queryset=StockTemp.objects.all()
    form = StockIssueForm(request.POST or None)
    #delete=models.CharField(max_length=50, blank=True, null=True) 
    delete=models.TextField ( blank=True, null=True)
    delete_items=[]
    quantity_issue=[]
    issue_to1=[]
    update_form=[]
    save_update_form=[]
    iss=[]
    iss2=[]
    iss=[]
    item_unit_kind_issue=[]
    del item_unit_kind_issue[:]
    item_unit_kind_issue_green=[]
    del item_unit_kind_issue_green[:]
    proceed_with_form=[]
    its_dup=False 
    item_unit_kind_issue_update=[]
    del item_unit_kind_issue_update[:]
    quntity_update=[]   
    #item_unit_kind_issue_update.clear()
    #quntity_update.clear()    
    list0={"a":"a", "b":"b","c":"c"}
  
    the_choices={}
    tuple_convert = StockTemp.item_unit_kind_choice_issue
    
    for a,b in tuple_convert:
        the_choices.setdefault(a,[]).append(b)


    form_kind= StockIssueForm2(request.POST or None)
    alchol_choice = TemplateList.objects.all()
    #delete lines
    for instance in StockTemp.objects.all():
        if (instance.item_fattal_code_issue ==0):
            instance.delete()
        instance.DateIssue=datetime.now().date()
        
        instance.save()

    
    
    if request.method == 'POST' and not form['item_fattal_code_issue']==None:
        form=StockIssueForm(request.POST)
        f=request.POST.get('item_fattal_code_issue')
        print ("f= ",f)
       
        if form.is_valid():
            #form.save(commit=False)
            if StockTemp.formisok==True:
                print ("formisok= ", StockTemp.formisok)
                form.save()
              


                
                
            else:
                print("errrrrr")

            

            
               


            for instance2 in StockTemp.objects.all():
                for instance1 in Stock.objects.all():
                    if str(instance2.item_fattal_code_issue) == str(instance1.item_fattal_code):
                        # 
                        instance2.item_name_issue = instance1.item_name
                        
                        if instance2.item_unit_kind_issue ==None:
                            print ("only than update quantity kind:  ", instance2.item_unit_kind_issue)
                            instance2.item_unit_kind_issue = instance1.item_unit_kind
                        #print (instance2.item_fattal_code_issue," found ", instance2.item_name_issue )
                        instance2.formisok=True
                        instance2.save()
           
        else:
            its_dup= True
    queryset=StockTemp.objects.all()
    
    context={
        "header": header,
        "queryset":queryset,
        "form": form,
        "delete":delete,
        "delete_items":delete_items,
        "quantity_issue":quantity_issue,
        "issue_to1":issue_to1,
        "item_unit_kind_issue":item_unit_kind_issue,
        "item_unit_kind_issue_green":item_unit_kind_issue_green,
        "form_kind":form_kind,
        "save_update_form":save_update_form,
        "update_form":update_form,
        "proceed_with_form":proceed_with_form,
        "its_dup":its_dup,
        "alchol_choice": alchol_choice,
        "the_choices":the_choices,
        "list0":list0,
 
    }
   
    quntity_update=request.POST.getlist("quantity_issue")
  
    issue_to=request.POST.get("issue_to1")
    issue_to_whom=request.POST.get("issue_to_whom")
    item_unit_kind_issue_update=request.POST.getlist("item_unit_kind_issue")
    item_unit_kind_issue_green=request.POST.getlist("item_unit_kind_issue_green")
    

    delete_items=request.POST.getlist("delete_items")
    # print (len(delete_items),"delete_items= ", delete_items )
    delete_checked_items(*delete_items)

    
    update_the_form_check = False
    update_main_db_check =False
    

    print ("\nBefore ****   function   start")
    print ("item_fattal_code_issue= ",item_unit_kind_issue )
    print ("issue_to= ",issue_to )
    print ("issue_to_whom= ",issue_to_whom )
    print ("item_unit_kind_issue", item_unit_kind_issue)
    print ("item_unit_kind_issue GREEN", item_unit_kind_issue_green)
    print ("item_unit_kind_issue update= ",item_unit_kind_issue_update )
    print ("quantity update= ", len(quntity_update) ," ", quntity_update)
    print ("**** ")
    print ("@@@@")
    for i in StockTemp.objects.all():
        print (i.id, end=' ')
        print (i.issue_quantity_transfer, end=' ')
        print (i.item_unit_kind_issue, end=' ')
        print (i.item_fattal_code_issue)
       
    print ("Before  @@@@  function, end")
    
    
    
    
    if (request.POST.get("update_main_db") != None):
        update_main_db_check = True
    else:
        update_main_db_check =False

    if (request.POST.get("update_the_form") != None):
        update_the_form_check = True
        print ("before function in the IFFFF ", update_the_form_check, update_main_db_check, "\n", item_unit_kind_issue_update)
        update_form_def (update_the_form_check, update_main_db_check, quntity_update, item_unit_kind_issue_update)



    clear_list(item_unit_kind_issue_update)
    clear_list(quntity_update)

    # print ("\n****  After function")
    # print ("issue_to= ",issue_to )
    # print ("issue_to_whom= ",issue_to_whom )
    # print ("item_unit_kind_issue", item_unit_kind_issue)
    # print ("item_unit_kind_issue update= ",item_unit_kind_issue_update )
    # print ("quantity update= ", len(quntity_update) ," ", quntity_update)
    # print ("**** ")
    # print ("@@@@")
    # for i in StockTemp.objects.all():
    #     print (i.id, end=' ')
    #     print (i.issue_quantity_transfer, end=' ')
    #     print (i.item_unit_kind_issue)
       
    # print ("@@@@ After function")
    
    
    
    #t=request.POST.get(all)
    #print("q post: \n",q)
    #print ("del form presses",z, len(z) )
    #q=request.GET.getlist("quantity_issue")
    #print("q get: \n",q)
    #print ("t kind ",t)
    #z=request.GET.getlist('delete_item')
 
    #reset DELETE the data of the form from DB 
    
    reset_form=  request.POST.get('delete')
    print("reset_form ",reset_form)
    if reset_form =="D" or reset_form=="DELETE" or reset_form =="d":
        for instance in StockTemp.objects.all():
            instance.delete()


    
    # print ("how many recordes: " , len(StockTemp.objects.all()))

    # print ("---choices")
    # tt=StockTemp.item_unit_kind_choice_issue
    # f= dict(tt)
    # print (f)
    # print (type(f))
    # if "grams" in f:
    #     print ("exist")  

    # for a,b in tt:
    #     if b == "grams":
    #         print ("wonderful")


    return render(request, 'index_issue.html',context)   



def list_items_view(request):
    header ='List of items'
    searchtitle= "Search Item"
    form = StockSearchForm(request.POST or None)
    queryset = Stock.objects.all()
    context ={
        "header": header,
        "queryset":queryset,
        "form": form,
                          
    }
    if request.method == 'POST':
        queryset = Stock.objects.filter(#item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         category_name__icontains=form['category_name'].value(),
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
           
    if request.method == 'POST':
        queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         #category_name__icontains=form['category_name'].value(),
                                         #item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
      
    context ={
        "form": form,
        "header": header,
        "queryset": queryset,
        "searchtitle": searchtitle
      
                
    }
    print("listview:")
    for instance in Stock.objects.all():
        print (instance.item_fattal_code, " issu:", instance.item_fattal_code_issue)
    print("end listview")
    return render(request, "list_items.html", context)




#EVERYTHING in the comment is also work also in the htm
def test_forms1(request):
    cars=models.CharField(max_length=50, blank=True, null=True) 

    dic1=[]
    vehicle=[]
    date= models.DateField()
    title=models.CharField(max_length=50, blank=True, null=True) 
    
    
 
  
    print (request.GET)
    input2= request.GET.getlist('vehicle')
    title= request.GET.getlist('title')
    date_input=request.GET.getlist('date')
    print ("input2= ", input2)
    print ("title= ", title)
    print ("date= ",date_input)
    if request.GET:
        input1=request.GET.getlist('cars')
        #input2=input1[0]
        #print(len(input1))
        print ("input1: ",input1)
        
        for instance in range(len(input1)):
            print("cars #", instance, " =",input1[instance])

        
        for instance in input1:
            print(instance)
        
        for instance in input2:
            print (instance)
        
        print ("len=", len(input1))
        if input1=='saab' or input1=='opel':
            return HttpResponse ("<h1>Great Choice </h1>")
        else:
            
            return redirect("home")

    
    #print(input2)
    #input1=input2.
    #print (input1)
    title='a'
    context={
        'title':title,
        'dic1':dic1,
        'cars':cars,
        'vehicle':vehicle,
        'date':date
      
       

    }

    return render(request, "testforms.html",context)







def StockFSearch_view(request):
    header ='Full Item Search'
    form = StockFullSearchForm(request.POST or None)
    queryset = Stock.objects.all()

    if request.method == 'POST':
        queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
  
    context ={
        "header": header,
        "queryset":queryset,
        "form": form,
                          
    }

        
    # if request.method == 'POST':
    #         # z=  form['item_name'].value()
    #         # zz= Stock.objects.values('item_name')
    #         # var2 = Stock.objects.values_list('item_name')
    #         # print ("z: ",z, ", var2:", var2, "zz: ",zz)
    #         # for i in var2:
    #         #     print (i)
    #         #     if z in i:
    #         #         print ("found ",z , " *****")
            
    #         # if isnumeric(form['quantity_min'].value())== False:
    #         #     print ("missing a number ***")
                
                
    #         queryset = Stock.objects.filter(category_name__icontains=form['category_name'].value(),
    #                                         item_name__icontains=form['item_name'].value(),
    #                                         quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value() ),
    #                                         item_fattal_code__icontains=form['item_fattal_code'].value(),
    #                                         description__icontains=form['description'].value(),
                                        
                                        
             #)
    if request.method == 'POST':
        queryset = Stock.objects.filter(#item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         category_name__icontains=form['category_name'].value(),
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
           
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
    #                                      #category_name__icontains=form['category_name'].value(),
    #                                      #item_name__iexact=form['item_name'].value(),
    #    )
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(item_fattal_code__range=(form['item_fattal_code_begin'].value(),form['item_fattal_code_end'].value()),
                                                     
        
    #     )
    
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value()),
                                                     
        
    #     )
    
    
    if request.method == 'POST':
        queryset = Stock.objects.filter(quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value()),
                                    item_fattal_code__range=(form['item_fattal_code_begin'].value(),form['item_fattal_code_end'].value())                                       
                                                     
        
        )
    
  
       
    context ={
        "form": form,
        "header": header,
        "queryset": queryset,
                
    }
    
    
    return render(request, "full_search.html", context)


def import_excel_view(request):


    form=import_excel( request.FILES)
    error_upload=''
    error=''
    if 'GET'== request.method: 
       return render(request,'excel_upload.html', {})
    else:
        try:
            excel_file = request.FILES["file"]
            print ("### \n",excel_file)
            
        except:
            
            excel_file = 'None'
            print ("### \n",excel_file)
            #messages.error(request, "This is not a valid file format (only csv or excel xlsx")
            #return HttpResponse("boom")
            error_upload="You can't  upload blank, only an EXCEL file" 
            context={
                "error_upload":error_upload

            }
            return render(request,'excel_upload.html', context)
    
    if not excel_file.name.endswith('.xlsx'):
        #error="\"" +excel_file.name + "\" is not a valid File"
        error= "This is not a valid File"
        error_file="\"" +excel_file.name +"\""
        context={

            "error":error,
            "error_file":error_file

        }

        return render(request,'excel_upload.html', context)
    else:
     # you may put validations here to check extension or file size
        print ("before wb ")
        wb = openpyxl.load_workbook(excel_file)

        #getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        excel_data = list()
        # print (worksheet.iter_rows())
        # print ("****** rows: ", worksheet.max_row) 
        # print ("****** col: ", worksheet.max_column) 
             
        #resetting values: 
        # first _run to skip the first row in ecel
        # skip+sskip update database if found new data
        # The validation done in differend python code: clean_import_files
        skip=False
        sskip=False
        row_counter=0
        first_run = True
        #print (row_counter)
        #print (row_data)
        # iterating over the rows and
        # getting value from each cell in row
        for  row in worksheet.iter_rows(min_row=1):
            row_data = list()
            #print ("top=",row_counter)
            for cell in row:
                row_data.append(str(cell.value))
            
            excel_data.append(row_data)
        # item_fattal_code_check = column[0]
        # clean_excel_item_fattal_code(item_fattal_code_check,skip)

        # item_name_check= column[1]
        # clean_excel_item_name(item_name_check,skip)
            
        # sup_name= column[2]
        # clean_excel_supplliers_name(sup_name,sskip)
            #print ("row counter: ", row_counter)
            skip=False
            sskip=False
        #print (row_counter)
        #print (row_data)
              
        
                
        
        
            for nonedata in range(row_counter):
                
                if row_data[nonedata] =='None':
                    row_data[nonedata]=''
                    #print ("found none")
            
            #before clean_excel_item_fattal_code 0   
            if not row_data[row_counter] =='None' and (first_run == False) :
                #before clean_excel_item_fattal_code 
                if row_data[row_counter].isnumeric:
                    item_fattal_code_check = row_data[row_counter]
                    clean_excel_item_fattal_code(item_fattal_code_check,skip)
                    #print (skip, " num ", row_data[row_counter], "=item fattal code ", row_counter, "=row counter" )
                #before clean_excel_item_name  1
                if row_data[row_counter + 1].isalpha:
                    item_name_check= row_data[row_counter+1]
                    #print ("before calling function= ", skip, " item name= ", item_name_check)
                    clean_excel_item_name(item_name_check,skip)
                    #print (skip, " after fiunction ", row_data[row_counter], "=item name ", row_counter, "=row counter" )
                #before clean_excel_sub_category_name   2            
                if (row_data[row_counter + 2]).isalnum:
                    sub_category_name= row_data[row_counter+2]
                    clean_excel_sub_category_name(sub_category_name,skip)
                    #print (skip, " sub+_category ", row_data[row_counter+2], "=sub category ", row_counter, "=row counter" )   
                    #print (skip, " alpha")
                
                # # #before suplier name   3
                # if (row_data[row_counter + 3]).isalnum:
                #     supplliers_name= row_data[row_counter+3]
                #     clean_excel_suppllers_name(supplliers_namek,skip)
                    #print (skip, " alpha ", row_data[row_counter], "=item name ", row_counter, "=row counter" )   
                    #print (skip, " alpha")
                #before suppliers_fattal_code  4 - no need to be clean. can be duplicate in same row input
            # if (row_data[row_counter + 4].isalnum:
                #   suppliers_fattal_code= row_data[row_counter+3]
                    #clean_excel_suppliers_fattal_code(suppliers_fattal_code,skip)
                    #print (skip, " alpha ", row_data[row_counter], "=item name ", row_counter, "=row counter" )   
                    #print (skip, " alpha")
            
                
                
                if skip == False:
                    print(type(row_data[row_counter]), " ", row_data[row_counter])
                    item_fattal_code_converted = int(row_data[row_counter])
                    _, created = Stock.objects.update_or_create(
                        item_fattal_code=item_fattal_code_converted,
                        item_name= row_data[row_counter+1],
                        sub_category_name=row_data[row_counter+2],
                        suppliers_fattal_code= row_data[row_counter+4],
                        category_name='',
                        description='',
                        
                    )
                if sskip == False:
                    _, created = SupplierInformation.objects.update_or_create(
                        suppliers_name=row_data[row_counter+3]

                ) 
                if skip == True:
                    skip = False
                if sskip == True:
                    sskip = False
            first_run = False
            row_counter = 0            
            

        #print(excel_data[1])
        #print (excel_data[1][4])
        # reading a cell
        #print(worksheet["A1"].value)
    
   
    
    #print ("form= ",form)
    #print (excel_file)
    context={
        "form":form,


    }

    #print ("not fell in upload ")
    return render(request, "excel_upload.html", context)






# def import_excel_file(request):
#     prompt = {
#         'order': 'MAKE SURE that this is csv or xlsx file, and the fields are in correct order'

#     }
    
#     if request.method == "GET":
#         file_data= request.FILES['file']
#         #if file_data.is_valid:
#         # return render (request, 'excel_upload1.html', {})
#     else:
#         return redirect ('/')
#         file_data= request.FILES['file']
#     if not file_data.name.endswith('.xlsx'):
#         messages.error(request, "This is not a valid file format (only csv or excel xlsx")
#         return redirect ('/')
#     else:
#     # you may put validations here to check extension or file size
        
#         wb = openpyxl.load_workbook(file_data)

#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#         print(worksheet)

#         # getting all sheets
#         sheets = wb.sheetnames
#         print(sheets)

#         # getting active sheet
#         active_sheet = wb.active
#         print(active_sheet)
        
#         excel_data = list()
#         print (worksheet.iter_rows())
#         print ("****** rows: ", worksheet.max_row) 
#         print ("****** col: ", worksheet.max_column) 
        

#         #print("number of columns: ", worksheet.iter_cols())       
        
#         # _, created = Stock.objects.update_or_create(
#         #             item_fattal_code=worksheet["A2"].value,
#         #             item_name= worksheet["B2"].value,

#         #     )
       
#         # iterating over the rows and
#         # getting value from each cell in row
#         row_counter=0
#         first_run = True
#         for  row in worksheet.iter_rows(min_row=1):
#             row_data = list()
#             print ("top=",row_counter)
        
#             for cell in row:
#                 row_data.append(str(cell.value))
                

#             excel_data.append(row_data)
#             # print(excel_data)
#             # print(len(excel_data))
#             # print(excel_data[0][0])
#             # #print (excel_data[row_counter])
#             # print ("****")
#             # print (len(row_data))    
#             # print (type(excel_data))
#             # print (len(excel_data))
#             # print (row_data[0])
#             # print (row_data[1])
#             # print("************")
            
#             #item_fattal_code_check = column[0]
#             #clean_excel_item_fattal_code(item_fattal_code_check,skip)

#             #item_name_check= column[1]
#             #clean_excel_item_name(item_name_check,skip)
            
#             #sup_name= column[2]
#             #clean_excel_supplliers_name(sup_name,sskip)
#             #print ("row counter: ", row_counter)
#             skip=False
#             sskip=False
#             print (row_counter)
#             print (row_data)
#             for nonedata in range(row_counter):
#                 if row_data[nonedata] =='None':
#                     row_data[nonedata]=''
#                     print ("found none")
#             #before clean_excel_item_fattal_code 0   
#             if not row_data[row_counter] =='None' and (first_run == False) :
#                 #before clean_excel_item_fattal_code 
#                 if row_data[row_counter].isnumeric:
#                     item_fattal_code_check = row_data[row_counter]
#                     clean_excel_item_fattal_code(item_fattal_code_check,skip)
#                     print (skip, " num ", row_data[row_counter], "=item fattal code ", row_counter, "=row counter" )

#                 #before clean_excel_item_name  1
#                 if row_data[row_counter + 1].isalpha:
#                     item_name_check= row_data[row_counter+1]
#                     print ("before calling function= ", skip, " item name= ", item_name_check)
#                     clean_excel_item_name(item_name_check,skip)
#                     print (skip, " after fiunction ", row_data[row_counter], "=item name ", row_counter, "=row counter" )
#                 #before clean_excel_sub_category_name   2            
#                 # if (row_data[row_counter + 2]).isalnum:
#                 #     sub_category_name= row_data[row_counter+2]
#                 #     clean_excel_sub_category_name(sub_category_name,skip)
#                 #     print (skip, " sub+_category ", row_data[row_counter+2], "=sub category ", row_counter, "=row counter" )   
#                 #     print (skip, " alpha")
                
#                 # #before suplier name   3
#                 # if (row_data[row_counter + 3].isalnum:
#                 #     supplliers_name= row_data[row_counter+3]
#                 #     clean_excel_supplliers_name(supplliers_namek,skip)
#                 #     #print (skip, " alpha ", row_data[row_counter], "=item name ", row_counter, "=row counter" )   
#                 #     #print (skip, " alpha")

#                 # #before suppliers_fattal_code  4 - no need to be clean. can be duplicate in same row input
#                 # if (row_data[row_counter + 4].isalnum:
#                 #     suppliers_fattal_code= row_data[row_counter+3]
#                 #     #clean_excel_suppliers_fattal_code(suppliers_fattal_code,skip)
#                 #     #print (skip, " alpha ", row_data[row_counter], "=item name ", row_counter, "=row counter" )   
#                 #     #print (skip, " alpha")
                
                
                
#                 if skip == False:
#                     print(type(row_data[row_counter]), " ", row_data[row_counter])
#                     item_fattal_code_converted = int(row_data[row_counter])
#                     _, created = Stock.objects.update_or_create(
#                         item_fattal_code=item_fattal_code_converted,
#                         item_name= row_data[row_counter+1],
#                         sub_category_name=row_data[row_counter+2],
#                         suppliers_fattal_code= row_data[row_counter+4],
#                         category_name='',
#                         description='',
                        

#                     )

#                 if sskip == False:
#                     _, created = SupplierInformation.objects.update_or_create(
#                         suppliers_name=row_data[row_counter+3]
 
#                    ) 
#                 if skip == True:
#                     skip = False
#                 if sskip == True:
#                     sskip = False
#             first_run = False
#             row_counter = 0            
            


        #print(excel_data[1])
        #print (excel_data[1][4])
        # reading a cell
        #print(worksheet["A1"].value)

        # return render(request, 'excel_upload1.html', {"excel_data":excel_data})





    


#@permission_required('admin.can_add_log_entry')


# def excel_import_example(request):
#     if "GET" == request.method:
#          return render(request, 'excel_upload.html', {})
#     else:
#         excel_file = request.FILES["excel_file"]

#         # you may put validations here to check extension or file size
#         if not excel_file.name.endswith('.xlsx'):
#             messages.error(request, "this is not excel  file")

#         wb = openpyxl.load_workbook(excel_file)

#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#         print(worksheet)

#         # getting all sheets
#         sheets = wb.sheetnames
#         print(sheets)

#         # getting active sheet
#         active_sheet = wb.active
#         print(active_sheet)

#         excel_data = list()
#         # iterating over the rows and
#         # getting value from each cell in row
#         for row in worksheet.iter_rows():
#             row_data = list()
#             for cell in row:
#                 row_data.append(str(cell.value))
#             excel_data.append(row_data)


#         # reading a cell
#         print(worksheet["A1"].value)

#     return render(request, 'excel_upload.html', {"excel_data":excel_data})




#***
# def import_excel_view(request):
#     form=import_excel( request.FILES)
#     error=''
#     error1=''
#     if 'GET'== request.method: 
#        return render(request,'excel_upload.html', {})
#     else:
#         try:
#             excel_file = request.FILES["file"]
#             print ("### \n",excel_file)
            
#         except:
            
#             excel_file = 'None'
#             print ("### \n",excel_file)
#             #messages.error(request, "This is not a valid file format (only csv or excel xlsx")
#             #return HttpResponse("boom")
#             error1="You must upload excel file"
#             context={
#                 "error1":error1,

#             }
#             return render(request,'excel_upload.html', context)
    
#     if not excel_file.name.endswith('.xlsx'):
#         error="This is not a valid File"
#         context={
#             "error":error,

#         }
#         return render(request,'excel_upload.html', context)
#     else:
#      # you may put validations here to check extension or file size
#         print ("before wb ")
#         wb = openpyxl.load_workbook(excel_file)

    
#     print ("form= ",form)
#     print (excel_file)
    
# #    #if request.method=='POST':
# #     #if not file_data.name.endswith('.xlsx'):   
    
        
        
      

#     context={
#         "form":form,


#     }

#     print ("not fell in upload ")
#     return render(request, "excel_upload.html", context)




    
def add_alchol_type(request):

    return render(request, "add_aclhol_type.html", {} )



    