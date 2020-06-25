import csv, io   #for csv
from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpRequest
from .forms import StockCreateForm, SuppliersCreateForm, StockSearchForm,StockUpdateForm, StockFullSearchForm
from .models import *
from .clean_import_files import *
from django.contrib import messages
import openpyxl
from datetime import datetime
from django.contrib.auth.decorators import permission_required  #for updateing from csv

# Create your views here.
 
 
def homepage_view(request):
    #this is the title for now
    #print (request.user)
    a=10
    b=7
    print (int(a / b))
    today = datetime.now().date()
    context= {
     
        "today": today,
        
    }

    return render(request,"home.html", context) 
"""  """


def delete_items(request,pk):
    #instance = get_object_or_404( id=pk)
    instance = Stock.objects.get(id=pk)
    instance.delete()
    return redirect('/list_items')

# 

def about_view(request):
    title='Welcome to my first Django project'
    #today = datetime.now().date()
    aboutmyself= "My name is Avi Sivan. My email onlyavi@gmail.com"
    context= {
        "title":title,
        "today": today,
        "aboutmyself":aboutmyself
        
    }
    
    return render(request, "about.html",context)


def under_construction(request):
   
    return render(request, "under_construction.html")

def add_newitem_view(request):
    form = StockCreateForm(request.POST or None)
    #z = form['item_name'].value()
    if form.is_valid():
        form.save()
        return redirect('/add_newitem')
    
    context ={
        "form":form,
        "title":"Add New Item",
        
    }
    return render(request, "add_newitem.html", context)

def update_items(request,pk):
    #pk is the object we want to update
    queryset= Stock.objects.get(id=pk)
    form=StockUpdateForm(instance=queryset)
    if request.method=='POST':
        form=StockUpdateForm(request.POST, instance=queryset)
        if form.is_valid():
            form.save()
            return redirect('/list_items')
    context={
        "form":form,
        
    }
    return render(request, 'add_newitem.html',context)       



def list_items_view(request):
    header ='List of items'
    searchtitle= "Search Item"
    form = StockSearchForm(request.POST or None)
    queryset = Stock.objects.all()
    context ={
        "header": header,
        "queryset":queryset,
        "form": form,
                          
    }
    if request.method == 'POST':
        queryset = Stock.objects.filter(#item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         category_name__icontains=form['category_name'].value(),
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
           
    if request.method == 'POST':
        queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         #category_name__icontains=form['category_name'].value(),
                                         #item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
      
    context ={
        "form": form,
        "header": header,
        "queryset": queryset,
        "searchtitle": searchtitle
      
                
    }
    
    return render(request, "list_items.html", context)



def StockFSearch_view(request):
    header ='Full Item Search'
    form = StockFullSearchForm(request.POST or None)
    queryset = Stock.objects.all()
  
    context ={
        "header": header,
        "queryset":queryset,
        "form": form,
                          
    }

        
    # if request.method == 'POST':
    #         # z=  form['item_name'].value()
    #         # zz= Stock.objects.values('item_name')
    #         # var2 = Stock.objects.values_list('item_name')
    #         # print ("z: ",z, ", var2:", var2, "zz: ",zz)
    #         # for i in var2:
    #         #     print (i)
    #         #     if z in i:
    #         #         print ("found ",z , " *****")
            
    #         # if isnumeric(form['quantity_min'].value())== False:
    #         #     print ("missing a number ***")
                
                
    #         queryset = Stock.objects.filter(category_name__icontains=form['category_name'].value(),
    #                                         item_name__icontains=form['item_name'].value(),
    #                                         quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value() ),
    #                                         item_fattal_code__icontains=form['item_fattal_code'].value(),
    #                                         description__icontains=form['description'].value(),
                                        
                                        
             #)
    if request.method == 'POST':
        queryset = Stock.objects.filter(#item_fattal_code__icontains=form['item_fattal_code'].value(),
                                         category_name__icontains=form['category_name'].value(),
                                         item_name__iexact=form['item_name'].value(),
                                                                                                                                                         
 
        )
    
           
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(item_fattal_code__icontains=form['item_fattal_code'].value(),
    #                                      #category_name__icontains=form['category_name'].value(),
    #                                      #item_name__iexact=form['item_name'].value(),
    #    )
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(item_fattal_code__range=(form['item_fattal_code_begin'].value(),form['item_fattal_code_end'].value()),
                                                     
        
    #     )
    
    # if request.method == 'POST':
    #     queryset = Stock.objects.filter(quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value()),
                                                     
        
    #     )
    
    
    if request.method == 'POST':
        queryset = Stock.objects.filter(quantity_item__range=(form['quantity_min'].value(),form['quantity_max'].value()),
                                    item_fattal_code__range=(form['item_fattal_code_begin'].value(),form['item_fattal_code_end'].value())                                       
                                                     
        
        )
    
  
       
    context ={
        "form": form,
        "header": header,
        "queryset": queryset,
                
    }
    
    
    return render(request, "full_search.html", context)


def csv_excel_file(request):
    prompt = {
        'order': 'MAKE SURE that this is csv or xlsx file, and the fields are in correct order'

    }
    if request.method == "GET":
        return render (request, 'excel_upload1.html', {})
    file_data= request.FILES['file']
    if not file_data.name.endswith('.csv') and not file_data.name.endswith('.xlsx'):
        messages.error(request, "This is not a valid file format (only csv or excel xlsx")
    elif file_data.name.endswith('.csv'):
        #csv_upload(request,file_data)
        csv_file =request.FILES['file']
        data_set = csv_file.read().decode('UTF-8')
        io_string = io.StringIO(data_set)
        next(io_string) #to skip the first line usually headrs
        for column in csv.reader(io_string, delimiter=',', quotechar="|" ):
            skip=False
            sskip=False
            item_fattal_code_check= column[0]
            clean_excel_item_fattal_code(item_fattal_code_check,skip)

            item_name_check= column[1]
            clean_excel_item_name(item_name_check,skip)
            
            # sup_name= column[2]
            # clean_excel_supplliers_name(sup_name,sskip)
    # item_name = a.cleaned_data.get('item_name')
        #category_name =self.cleaned_data.get('category_name')
        
            
        
            #clean_excel_item_fattal_code(item_name_check,skip)
            if skip == False:
                item_fattal_code_converted = int(item_fattal_code_check)
                # _, created = Stock.objects.update_or_create(
                #     item_fattal_code=item_fattal_code_check,
                #     item_name= item_name_check,

                # )
                _, created = Stock.objects.update_or_create(

                    item_fattal_code=item_fattal_code_converted,
                    item_name= item_name_check,
                    #sub_category_name=row_data[row_counter+2],
                    #suppliers_fattal_code= row_data[row_counter+4],
                    category_name='',
                    description='',
                )
            if sskip==False:
                _, created = SupplierInformation.objects.update_or_create(
                    suppliers_name=column[2]

                ) 
            if skip == True:
                skip = False
            if sskip == True:
                sskip = False
            
        context ={}
        #so it can be change in the future
        
        return render(request,'excel_upload1.html',context)

    #  *** EXCEL FILE INPUT CHECK ***

    elif file_data.name.endswith('.xlsx'):
        #excel_upload(request)
        #excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        
        wb = openpyxl.load_workbook(file_data)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        
        excel_data = list()
        print (worksheet.iter_rows())
        print ("****** rows: ", worksheet.max_row) 
        print ("****** col: ", worksheet.max_column) 
        

        #print("number of columns: ", worksheet.iter_cols())       
        
        # _, created = Stock.objects.update_or_create(
        #             item_fattal_code=worksheet["A2"].value,
        #             item_name= worksheet["B2"].value,

        #     )
       
        # iterating over the rows and
        # getting value from each cell in row
        row_counter=0
        first_run = True
        for  row in worksheet.iter_rows(min_row=1):
            row_data = list()
            print ("top=",row_counter)
        
            for cell in row:
                row_data.append(str(cell.value))
                

            excel_data.append(row_data)
            # print(excel_data)
            # print(len(excel_data))
            # print(excel_data[0][0])
            # #print (excel_data[row_counter])
            # print ("****")
            # print (len(row_data))    
            # print (type(excel_data))
            # print (len(excel_data))
            # print (row_data[0])
            # print (row_data[1])
            # print("************")
            
            #item_fattal_code_check = column[0]
            #clean_excel_item_fattal_code(item_fattal_code_check,skip)

            #item_name_check= column[1]
            #clean_excel_item_name(item_name_check,skip)
            
            #sup_name= column[2]
            #clean_excel_supplliers_name(sup_name,sskip)
            #print ("row counter: ", row_counter)
            skip=False
            sskip=False
            print (row_counter)
            print (row_data)
            for nonedata in range(row_counter):
                if row_data[nonedata] =='None':
                    row_data[nonedata]=''
                    print ("found none")
            #before clean_excel_item_fattal_code 0   
            if not row_data[row_counter] =='None' and (first_run == False) :
                #before clean_excel_item_fattal_code 
                if row_data[row_counter].isnumeric:
                    item_fattal_code_check = row_data[row_counter]
                    clean_excel_item_fattal_code(item_fattal_code_check,skip)
                    print (skip, " num ", row_data[row_counter], "=item fattal code ", row_counter, "=row counter" )

                #before clean_excel_item_name  1
                if row_data[row_counter + 1].isalpha:
                    item_name_check= row_data[row_counter+1]
                    print ("before calling function= ", skip, " item name= ", item_name_check)
                    clean_excel_item_name(item_name_check,skip)
                    print (skip, " after fiunction ", row_data[row_counter], "=item name ", row_counter, "=row counter" )
                #before clean_excel_sub_category_name   2            
                # if (row_data[row_counter + 2]).isalnum:
                #     sub_category_name= row_data[row_counter+2]
                #     clean_excel_sub_category_name(sub_category_name,skip)
                #     print (skip, " sub+_category ", row_data[row_counter+2], "=sub category ", row_counter, "=row counter" )   
                #     print (skip, " alpha")
                
                # #before suplier name   3
                # if (row_data[row_counter + 3].isalnum:
                #     supplliers_name= row_data[row_counter+3]
                #     clean_excel_supplliers_name(supplliers_namek,skip)
                #     #print (skip, " alpha ", row_data[row_counter], "=item name ", row_counter, "=row counter" )   
                #     #print (skip, " alpha")

                # #before suppliers_fattal_code  4 - no need to be clean. can be duplicate in same row input
                # if (row_data[row_counter + 4].isalnum:
                #     suppliers_fattal_code= row_data[row_counter+3]
                #     #clean_excel_suppliers_fattal_code(suppliers_fattal_code,skip)
                #     #print (skip, " alpha ", row_data[row_counter], "=item name ", row_counter, "=row counter" )   
                #     #print (skip, " alpha")
                
                
                
                if skip == False:
                    print(type(row_data[row_counter]), " ", row_data[row_counter])
                    item_fattal_code_converted = int(row_data[row_counter])
                    _, created = Stock.objects.update_or_create(
                        item_fattal_code=item_fattal_code_converted,
                        item_name= row_data[row_counter+1],
                        sub_category_name=row_data[row_counter+2],
                        suppliers_fattal_code= row_data[row_counter+4],
                        category_name='',
                        description='',
                        

                    )

                if sskip == False:
                    _, created = SupplierInformation.objects.update_or_create(
                        suppliers_name=row_data[row_counter+3]
 
                   ) 
                if skip == True:
                    skip = False
                if sskip == True:
                    sskip = False
            first_run = False
            row_counter = 0            
            


        #print(excel_data[1])
        #print (excel_data[1][4])
        # reading a cell
        #print(worksheet["A1"].value)

        return render(request, 'excel_upload1.html', {"excel_data":excel_data})
    




    


#@permission_required('admin.can_add_log_entry')


# def excel_import_example(request):
#     if "GET" == request.method:
#          return render(request, 'excel_upload.html', {})
#     else:
#         excel_file = request.FILES["excel_file"]

#         # you may put validations here to check extension or file size
#         if not excel_file.name.endswith('.xlsx'):
#             messages.error(request, "this is not excel  file")

#         wb = openpyxl.load_workbook(excel_file)

#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#         print(worksheet)

#         # getting all sheets
#         sheets = wb.sheetnames
#         print(sheets)

#         # getting active sheet
#         active_sheet = wb.active
#         print(active_sheet)

#         excel_data = list()
#         # iterating over the rows and
#         # getting value from each cell in row
#         for row in worksheet.iter_rows():
#             row_data = list()
#             for cell in row:
#                 row_data.append(str(cell.value))
#             excel_data.append(row_data)


#         # reading a cell
#         print(worksheet["A1"].value)

#     return render(request, 'excel_upload.html', {"excel_data":excel_data})
    


    